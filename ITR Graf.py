import openpyxl
import datetime
import re
import calendar
import locale

class Picker():

    def __init__(self):
        pass
    
    #Недописанный метод
    def GrafWriteInDate(self, id: int, date: datetime.date):
        book = openpyxl.load_workbook(r"C:\Users\rezon\OneDrive\Рабочий стол\MAIN_Grafik_ITR_2017 test.xlsx")
        sheet = book.active
        #startIdPosition = "O23"
        
        print(date.year, date.month, date.day)
        
        sheet['U27'] = "пр"
        print(sheet["U27"].value)
        book.save(r"C:\Users\rezon\OneDrive\Рабочий стол\MAIN_Grafik_ITR_2017 test.xlsx")
        book.close()
    def GrafWriteOutDate(self, id, Date):
        pass
    def GrafWriteWorkInterval(self, id, firstDate, secondDate):
        pass 
    def GrafWriteHollidayInterval(self, id, firstDate, secondDate):
        pass
    # Методы для перехода с прошлого года
    def GrafStartOut(self, id, Date):
        pass
    def GrafEndIn(self, id, Date):
        pass
    def GrafHoliDayStartOut(self, id, Date):
        pass
    #GrafHoliDayEndIn(Date) - Не надо делать тк лучше не делать переходный отпуск
    def dateColumnSearch(self, date: datetime.date) -> int:
        book = openpyxl.load_workbook(r"C:\Users\rezon\OneDrive\Рабочий стол\MAIN_Grafik_ITR_2017 test.xlsx")
        sheet = book.active
        column_letter_start = 0
        row_with_monthDay = 0
        search_date_column = 0
        
        locale.setlocale(locale.LC_TIME, 'ru_RU')  #Переключаемся на русский язык (переводит месяца(1-12) на русский язык для поиска)
        month = calendar.month_name[date.month]
    
        #Search month
        for row in sheet.rows:
            for cell in row:
                if re.match(month, str(cell.value)):
                    row_with_monthDay = cell.row + 1
                    column_letter_start = cell.column
                    
        #Search day in month
        for row_cells in sheet.iter_rows(min_row=row_with_monthDay, max_row=row_with_monthDay, min_col=column_letter_start, max_col=column_letter_start+31):
            for cell in row_cells:
                if re.match(str(date.day), str(cell.value)):
                    search_date_column = cell.column
                    break  
                      
        book.close()
        return search_date_column
   


Picker1 = Picker()
date1 = datetime.date(2017,1,22)
#Picker1.dateColumnSearch(date1)

print(Picker1.dateColumnSearch(date1))
print("Be happy")